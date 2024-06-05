import requests
import openpyxl 
from openpyxl.drawing.image import Image as openpyxl_Image
from bs4 import BeautifulSoup as bs 
import re 

#for the image module, need requests also 
from PIL import Image
from io import BytesIO

def create_img_object_from_link(*, url_link_args):
    response = requests.get(url_link_args)
    img = Image.open(BytesIO(response.content))
    img_1 = openpyxl_Image(img)
    return img_1
    
##
#   @mainpage Doxygen Documentation for NTUSU food recommendation 
#   @section main_description Description
#   This projects aims to scrap restaurant data from website and put it into a ExcelSheet 
#   @section main_note Notes 
#   N/A 

##
#   @brief Defines the code for web scrapping 
#
#   @section description_web_scraping_NTUSU Description 
#   This is a web scrapping project to scrap the best eats in Singapore
#
#   @section note_web_scraping_NTUSU_description Notes 
#   This is the code section
#
# 

website_url = "https://www.timeout.com/singapore/restaurants/best-cheap-eats-in-singapore"
excel_path  = r"C:\Users\Ng Hong Xi\Desktop\Food_recommendation_for_NTUSU.xlsx"

website_html_img_tag_array = []   
excel_row_index = 2 

#image array for storing new image object 
image_link_array = [] 
image_object_array = []

def main():
    """!This is the entry point     
    """
    #open text file 
    text_file = open("logs.txt", "w")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    global excel_row_index  # Declare excel_row_index as global

    website = requests.get(website_url)
    website_html = bs(website.text, "html.parser")
    website_html_div_tag_array = website_html.find_all('div', attrs={'class': '_imageWrap_vapn8_242'})        
    #this is to get <img> tag from <div> tag 
    for i in range (len(website_html_div_tag_array)):
        website_html_img_tag_array.append(website_html_div_tag_array[i].picture.img)        
            
    #this is to fill in the first row inside excel   
    fill_excel_row_title(worksheet_args = ws)
    #end
    
    #ws.add_image(image_object_2, "C2")
     
    for i in range (len(website_html_img_tag_array)):

        image_title = f"{find_ideal_image_size_http_title(img_tag_args = website_html_img_tag_array[i])}"
        ws[f"A{excel_row_index}"] = image_title

        image_http_link = f"{find_ideal_image_size_http_link(srcset_tag_args = srcset_tag_get(website_html_image_tag_args = website_html_img_tag_array[i]))}"
        ws[f"B{excel_row_index}"] = image_http_link

        #store the image link inside image_link_array
        image_link_array.append(image_http_link) 

        #shutdown
        image_object_array.append(create_img_object_from_link(url_link_args = image_link_array[i]))
        ws.add_image(image_object_array[i], f"C{excel_row_index}")
        
        excel_row_index += 1
        text_file.write("\n")
    
    wb.save(excel_path)

#working correctly 
def srcset_tag_get(*, website_html_image_tag_args = "None"):
    '''!This gets the srcset string used in lazy loading

    @param website_html_image_tag_args This is the individual <img> tag for the html 
    @return Tag - The value of the srcset attribute
     '''
    website_image_tag_srcset_attrs = website_html_image_tag_args["data-srcset"]
    return website_image_tag_srcset_attrs

#sets 380 as default 
def find_ideal_image_size_http_link(*, required_width = "380", srcset_tag_args = "None"):
    '''!This gets the exact image http link

    @param required_size This is the width of the image
    @param srcset_tag_args This is the srcset tag 
    @return String - The http link of the image 
    '''
    srcset_string = str(srcset_tag_args)
    #print(srcset_string)    
    pattern = r"(https://[^\s]*" + required_width + "[^\s]*)"

    #returns the Match object of this re
    search_status = re.search(pattern, srcset_string)
    if search_status is not None:
        #this obtains the first image of 380w
        #print(search_status.group(0))
        return (search_status.group(0))
    
 

    
        

def find_ideal_image_size_http_title(*, img_tag_args = "None"):
    '''!This gets the name of the image through the title attribute for the <img> tag 
    @param website_html_image_tag This is the <img> tag 
    @return String - The name of the image 
    '''
    return (str(img_tag_args["title"]))
    
    
def fill_excel_row_title(*, worksheet_args = "None"):
    '''!This fills up the first row of the excel sheet 
    @param worksheet_args This is the worksheet object 
    '''
    title = ["restaurant name", "image link"] 
    for rows in range(len(title)): 
        worksheet_args.cell(row = 1, column = rows + 1, value = title[rows])



# image_link = "https://media.timeout.com/images/106080273/380/285/image.jpg"
# image_link_2 = "https://media.timeout.com/images/105913471/380/285/image.jpg"
# image_object = create_img_object_from_link(url_link_args = image_link)
# image_object_2 = create_img_object_from_link(url_link_args = image_link_2)
main()


